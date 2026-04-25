"""
NigeriaAgriScope — Module 6: Power BI Data Mart Export
=======================================================
Reads all upstream module outputs (M1–M5) and writes a clean, denormalised,
Power BI-ready Excel workbook with 8 sheets in a star-schema layout.

Star schema design
------------------
  Hub (dimension) : dim_zone_crop         — 42 rows, zone × crop
  Fact tables     : fact_yield_history    — historical yield/climate/fertilizer
                    fact_yield_forecast   — XGBoost test set predictions (M4)
                    fact_production_forecast — Prophet 2024–2026 (M4)
                    fact_input_requirements  — fertilizer + cost per ha (M5)
                    fact_planting_calendar   — planting/harvest windows (M5)
                    fact_operations_summary  — key milestone months (M5)
                    fact_operations_detail   — full week-by-week schedule (M5)

All fact tables join to dim_zone_crop on (zone, crop) composite key.
fact_operations_detail joins on (zone, crop) with drill-through to summary.

Power BI import: see README_PowerBI.md in the same output folder.

Output
------
  module6_powerbi/outputs/NigeriaAgriScope_PowerBI_DataMart.xlsx
  module6_powerbi/outputs/README_PowerBI.md

Author : Fidelis Akinbule
Date   : April 2026
"""

from __future__ import annotations

import logging
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("m6_powerbi_export")

# ── Paths ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent

SOURCES = {
    "master_table": ROOT / "data" / "processed" / "master_table.csv",
    "yield_predictions": ROOT / "module4_models" / "outputs" / "yield_predictions.csv",
    "forecast_palm": ROOT
    / "module4_models"
    / "outputs"
    / "production_forecast_oil_palm_fruit.csv",
    "forecast_cassava": ROOT
    / "module4_models"
    / "outputs"
    / "production_forecast_cassava.csv",
    "forecast_maize": ROOT
    / "module4_models"
    / "outputs"
    / "production_forecast_maize.csv",
    "input_requirements": ROOT
    / "module5_planning"
    / "outputs"
    / "input_requirements_enhanced.csv",
    "planting_calendar": ROOT
    / "module5_planning"
    / "outputs"
    / "planting_calendar_all_zones.csv",
    "operations_summary": ROOT
    / "module5_planning"
    / "outputs"
    / "operations_schedule_summary.csv",
    "operations_detail": ROOT
    / "module5_planning"
    / "outputs"
    / "operations_schedule_all_zones.csv",
}

OUT_DIR = ROOT / "module6_powerbi" / "outputs"
OUT_DIR.mkdir(parents=True, exist_ok=True)
WORKBOOK_PATH = OUT_DIR / "NigeriaAgriScope_PowerBI_DataMart.xlsx"
README_PATH = OUT_DIR / "README_PowerBI.md"

# ── Colour palette (professional, Nigeria-themed) ──────────────────────────
CLR_HEADER_DARK = "1F4E79"  # dark navy — dimension table headers
CLR_HEADER_MID = "2E75B6"  # mid blue  — fact table headers
CLR_HEADER_GREEN = "375623"  # dark green — operations/planning headers
CLR_HEADER_BROWN = "7B3F00"  # brown      — cost/input headers
CLR_ALT_ROW = "EBF3FB"  # light blue — alternating row fill
CLR_WHITE = "FFFFFF"
CLR_FONT_LIGHT = "FFFFFF"
CLR_FONT_DARK = "1F1F1F"

MONTH_NAMES = [
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
]


# ── Dimension lookups (built from constants, not from data) ────────────────

ZONE_TYPE = {
    "North West": "North",
    "North East": "North",
    "North Central": "North",
    "South West": "South",
    "South East": "South",
    "South South": "South",
}

ZONE_RAINFALL_CLASS = {
    "North West": "Sudan Savanna",
    "North East": "Sahel",
    "North Central": "Guinea Savanna",
    "South West": "Rainforest/Derived Savanna",
    "South East": "Rainforest",
    "South South": "Mangrove/Rainforest",
}

CROP_CATEGORY = {
    "Cassava": "Root & Tuber",
    "Yam": "Root & Tuber",
    "Maize": "Cereal",
    "Rice (paddy)": "Cereal",
    "Sorghum": "Cereal",
    "Oil palm fruit": "Tree Crop",
    "Cocoa beans": "Tree Crop",
}

CROP_PERENNIAL = {
    "Cassava": False,
    "Yam": False,
    "Maize": False,
    "Rice (paddy)": False,
    "Sorghum": False,
    "Oil palm fruit": True,
    "Cocoa beans": True,
}

STATE_GROUPS = {
    "North West": "Kano, Kaduna, Sokoto, Katsina, Zamfara, Kebbi, Jigawa",
    "North East": "Borno, Yobe, Adamawa, Gombe, Bauchi, Taraba",
    "North Central": "Benue, Niger, Plateau, Kogi, Kwara, Nasarawa, FCT",
    "South West": "Lagos, Ogun, Oyo, Osun, Ondo, Ekiti",
    "South East": "Anambra, Imo, Abia, Enugu, Ebonyi",
    "South South": "Rivers, Delta, Edo, Bayelsa, Cross River, Akwa Ibom",
}


# ── Source validation ──────────────────────────────────────────────────────


def validate_sources() -> None:
    """Raise FileNotFoundError with actionable message if any source missing."""
    module_map = {
        "master_table": "Module 1 (generate_data.py)",
        "yield_predictions": "Module 4 (yield_model.py)",
        "forecast_palm": "Module 4 (production_forecast.py)",
        "forecast_cassava": "Module 4 (production_forecast.py)",
        "forecast_maize": "Module 4 (production_forecast.py)",
        "input_requirements": "Module 5 (input_calculator.py)",
        "planting_calendar": "Module 5 (planting_calendar.py)",
        "operations_summary": "Module 5 (operations_schedule.py)",
        "operations_detail": "Module 5 (operations_schedule.py)",
    }
    missing = []
    for key, path in SOURCES.items():
        if not path.exists():
            missing.append(f"  {path.name}  →  run {module_map[key]}")
    if missing:
        raise FileNotFoundError("Missing upstream files:\n" + "\n".join(missing))
    log.info("Source validation passed — all 9 upstream files present")


# ── Table builders ─────────────────────────────────────────────────────────


def build_dim_zone_crop() -> pd.DataFrame:
    """Sheet 1 — 42-row master dimension table."""
    zones = list(ZONE_TYPE.keys())
    crops = list(CROP_CATEGORY.keys())
    rows = []
    for zone in zones:
        for crop in crops:
            rows.append(
                {
                    "zone": zone,
                    "crop": crop,
                    "state_group": STATE_GROUPS[zone],
                    "zone_type": ZONE_TYPE[zone],
                    "zone_rainfall_class": ZONE_RAINFALL_CLASS[zone],
                    "crop_category": CROP_CATEGORY[crop],
                    "is_perennial": CROP_PERENNIAL[crop],
                    "zone_crop_key": f"{zone}|{crop}",
                }
            )
    df = pd.DataFrame(rows)
    log.info("dim_zone_crop: %d rows × %d cols", *df.shape)
    return df


def build_fact_yield_history() -> pd.DataFrame:
    """Sheet 2 — Historical yield, climate, fertilizer per zone-crop-year."""
    df = pd.read_csv(SOURCES["master_table"])
    df["year"] = df["year"].astype(int)

    keep = [
        "zone",
        "crop",
        "year",
        "area_ha",
        "production_tonnes",
        "yield_hg_ha",
        "rainfall_mm_annual",
        "temp_avg_celsius",
        "humidity_pct",
        "solar_radiation",
        "fertilizer_kg_ha",
        "wb_fertilizer_kg_ha",
        "agric_gdp_share",
        "rural_population",
    ]
    keep = [c for c in keep if c in df.columns]
    df = df[keep].copy()

    # Derived column: yield in tonnes/ha (more intuitive than hg/ha for PBI)
    df["yield_tonnes_ha"] = (df["yield_hg_ha"] / 10_000).round(4)

    # Composite key for PBI relationship
    df["zone_crop_key"] = df["zone"] + "|" + df["crop"]

    # Round numerics
    for col in [
        "area_ha",
        "production_tonnes",
        "rainfall_mm_annual",
        "temp_avg_celsius",
        "humidity_pct",
        "solar_radiation",
        "fertilizer_kg_ha",
        "wb_fertilizer_kg_ha",
    ]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").round(2)

    df = df.sort_values(["zone", "crop", "year"]).reset_index(drop=True)
    log.info("fact_yield_history: %d rows × %d cols", *df.shape)
    return df


def build_fact_yield_forecast() -> pd.DataFrame:
    """Sheet 3 — XGBoost test set predictions (2019–2023)."""
    df = pd.read_csv(SOURCES["yield_predictions"])

    # Ensure all expected columns present
    expected = [
        "zone",
        "crop",
        "year",
        "yield_hg_ha",
        "predicted_yield_hg_ha",
        "residual_hg_ha",
        "abs_pct_error",
    ]
    present = [c for c in expected if c in df.columns]
    df = df[present].copy()

    df["year"] = df["year"].astype(int)
    df["yield_tonnes_ha_actual"] = (df["yield_hg_ha"] / 10_000).round(4)
    df["yield_tonnes_ha_predicted"] = (df["predicted_yield_hg_ha"] / 10_000).round(4)
    df["zone_crop_key"] = df["zone"] + "|" + df["crop"]
    df["data_type"] = "XGBoost Test Set (2019–2023)"

    df = df.sort_values(["zone", "crop", "year"]).reset_index(drop=True)
    log.info("fact_yield_forecast: %d rows × %d cols", *df.shape)
    return df


def build_fact_production_forecast() -> pd.DataFrame:
    """Sheet 4 — Prophet production forecasts for 3 crops (2000–2026)."""
    frames = []
    crop_map = {
        "forecast_palm": "Oil palm fruit",
        "forecast_cassava": "Cassava",
        "forecast_maize": "Maize",
    }
    for key, crop_name in crop_map.items():
        df = pd.read_csv(SOURCES[key])
        df["crop"] = crop_name
        frames.append(df)

    df = pd.concat(frames, ignore_index=True)
    df["year"] = pd.to_numeric(df["year"], errors="coerce").astype(int)

    # Ensure CI columns present (actual rows have NaN — keep as blank in Excel)
    for col in ["lower_80ci_tonnes", "upper_80ci_tonnes"]:
        if col not in df.columns:
            df[col] = None

    df["production_million_tonnes"] = (
        pd.to_numeric(df["production_tonnes"], errors="coerce") / 1_000_000
    ).round(3)

    col_order = [
        "crop",
        "year",
        "type",
        "production_tonnes",
        "production_million_tonnes",
        "lower_80ci_tonnes",
        "upper_80ci_tonnes",
    ]
    df = df[[c for c in col_order if c in df.columns]]
    df = df.sort_values(["crop", "year"]).reset_index(drop=True)
    log.info("fact_production_forecast: %d rows × %d cols", *df.shape)
    return df


def build_fact_input_requirements() -> pd.DataFrame:
    """Sheet 5 — Enhanced input requirements (fertilizer + seed + labour + cost)."""
    df = pd.read_csv(SOURCES["input_requirements"])

    # Drop M4 polynomial internals — not needed in Power BI
    drop_cols = [c for c in df.columns if c.startswith("poly_")]
    df = df.drop(columns=drop_cols, errors="ignore")

    # Drop low-signal reliability flag columns (keep score)
    df = df.drop(columns=["is_reliable"], errors="ignore")

    df["zone_crop_key"] = df["zone"] + "|" + df["crop"]

    # Month name for planting (join from calendar if present)
    numeric_cols = [
        "recommended_min_kg_ha",
        "recommended_max_kg_ha",
        "pdr_kg_ha",
        "current_avg_kg_ha",
        "yield_gap_at_pdr_hg_ha",
        "seed_cost_naira_per_ha",
        "labour_person_days_per_ha",
        "labour_cost_naira_per_ha",
        "agrochemical_cost_naira_per_ha",
        "fertilizer_cost_naira_per_ha",
        "total_input_cost_naira_per_ha",
        "total_input_cost_usd_per_ha",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").round(2)

    df = df.sort_values(["zone", "crop"]).reset_index(drop=True)
    log.info("fact_input_requirements: %d rows × %d cols", *df.shape)
    return df


def build_fact_planting_calendar() -> pd.DataFrame:
    """Sheet 6 — Planting windows and reliability per zone-crop."""
    df = pd.read_csv(SOURCES["planting_calendar"])

    # Truncate notes to 150 chars (Power BI tooltip practical limit)
    if "planting_notes" in df.columns:
        df["planting_notes"] = df["planting_notes"].str[:150]

    df["zone_crop_key"] = df["zone"] + "|" + df["crop"]

    # Add month number names for both planting and harvest (for PBI slicer)
    if "planting_month" in df.columns:
        df["planting_month_num"] = df["planting_month"].astype(int)
    if "harvest_month" in df.columns:
        df["harvest_month_num"] = df["harvest_month"].astype(int)

    df = df.sort_values(["zone", "crop"]).reset_index(drop=True)
    log.info("fact_planting_calendar: %d rows × %d cols", *df.shape)
    return df


def build_fact_operations_summary() -> pd.DataFrame:
    """Sheet 7 — Key milestone months per zone-crop (compact pivot)."""
    df = pd.read_csv(SOURCES["operations_summary"])
    df["zone_crop_key"] = df["zone"] + "|" + df["crop"]
    df = df.sort_values(["zone", "crop"]).reset_index(drop=True)
    log.info("fact_operations_summary: %d rows × %d cols", *df.shape)
    return df


def build_fact_operations_detail() -> pd.DataFrame:
    """Sheet 8 — Full week-by-week schedule (drill-through from summary)."""
    df = pd.read_csv(SOURCES["operations_detail"])

    # Keep only columns useful in Power BI (drop redundant month numbers
    # that would clutter the field list)
    keep = [
        "zone",
        "crop",
        "phase",
        "activity",
        "start_week_offset",
        "end_week_offset",
        "approx_start_month",
        "approx_end_month",
        "planting_month_name",
        "harvest_month_name",
        "rainfall_reliability_score",
        "activity_notes",
    ]
    keep = [c for c in keep if c in df.columns]
    df = df[keep].copy()

    # Truncate notes
    if "activity_notes" in df.columns:
        df["activity_notes"] = df["activity_notes"].str[:200]

    df["zone_crop_key"] = df["zone"] + "|" + df["crop"]

    # Phase sort order for correct timeline display in PBI
    phase_order = {
        "PRE_PLANTING": 1,
        "PLANTING": 2,
        "EARLY_GROWTH": 3,
        "VEGETATIVE": 4,
        "REPRODUCTIVE": 5,
        "PRE_HARVEST": 6,
        "HARVEST": 7,
        "POST_HARVEST": 8,
    }
    df["phase_sort"] = df["phase"].map(phase_order).fillna(9).astype(int)

    df = df.sort_values(
        ["zone", "crop", "phase_sort", "start_week_offset"]
    ).reset_index(drop=True)

    log.info("fact_operations_detail: %d rows × %d cols", *df.shape)
    return df


# ── Excel formatting ───────────────────────────────────────────────────────


def _make_header_fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_colour, fgColor=hex_colour)


def _make_header_font(bold: bool = True) -> Font:
    return Font(name="Arial", bold=bold, color=CLR_FONT_LIGHT, size=10)


def _make_data_font() -> Font:
    return Font(name="Arial", size=9, color=CLR_FONT_DARK)


def _make_alt_fill() -> PatternFill:
    return PatternFill("solid", start_color=CLR_ALT_ROW, fgColor=CLR_ALT_ROW)


def _thin_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(bottom=thin)


def format_sheet(
    ws,
    header_colour: str,
    freeze_row: int = 2,
    col_widths: dict | None = None,
) -> None:
    """
    Apply consistent formatting to a worksheet:
    - Header row: coloured fill, white bold Arial font, centre-aligned
    - Data rows: alternating light blue fill, left-aligned
    - Frozen header row
    - Auto-filter on header row
    - Column widths from col_widths dict or auto-sized
    """
    header_fill = _make_header_fill(header_colour)
    header_font = _make_header_font()
    data_font = _make_data_font()
    alt_fill = _make_alt_fill()

    max_col = ws.max_column
    max_row = ws.max_row

    # Format header row (row 1)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # Format data rows with alternating fill
    for row in range(2, max_row + 1):
        fill = alt_fill if row % 2 == 0 else None
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    # Freeze header
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Row height for header
    ws.row_dimensions[1].height = 32

    # Column widths
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    else:
        # Auto-size based on max content length (capped at 45)
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            max_len = 0
            for row in range(1, min(max_row + 1, 200)):  # sample first 200 rows
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)


# ── Workbook assembly ──────────────────────────────────────────────────────

SHEET_CONFIGS = [
    # (sheet_name, df_builder, header_colour, col_widths_override)
    ("dim_zone_crop", build_dim_zone_crop, CLR_HEADER_DARK, None),
    ("fact_yield_history", build_fact_yield_history, CLR_HEADER_MID, None),
    ("fact_yield_forecast", build_fact_yield_forecast, CLR_HEADER_MID, None),
    ("fact_production_forecast", build_fact_production_forecast, CLR_HEADER_MID, None),
    ("fact_input_requirements", build_fact_input_requirements, CLR_HEADER_BROWN, None),
    ("fact_planting_calendar", build_fact_planting_calendar, CLR_HEADER_GREEN, None),
    ("fact_operations_summary", build_fact_operations_summary, CLR_HEADER_GREEN, None),
    ("fact_operations_detail", build_fact_operations_detail, CLR_HEADER_GREEN, None),
]


def write_workbook(sheets: dict[str, pd.DataFrame]) -> None:
    """Write all DataFrames to Excel then apply openpyxl formatting."""

    # Step 1: write raw data via pandas ExcelWriter
    with pd.ExcelWriter(WORKBOOK_PATH, engine="openpyxl") as writer:
        for name, (df, _, _) in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
    log.info("Raw data written → %s", WORKBOOK_PATH.name)

    # Step 2: post-process with openpyxl for formatting
    wb = load_workbook(WORKBOOK_PATH)
    for name, (_, header_colour, col_widths) in sheets.items():
        ws = wb[name]
        format_sheet(ws, header_colour, col_widths=col_widths)
        log.info(
            "  Formatted %-30s  rows=%-5d cols=%d",
            name,
            ws.max_row - 1,
            ws.max_column,
        )

    # Add a cover sheet as the first tab
    cover = wb.create_sheet("_COVER", 0)
    _write_cover(cover)

    wb.save(WORKBOOK_PATH)
    log.info("Workbook saved → %s", WORKBOOK_PATH)


def _write_cover(ws) -> None:
    """Write a human-readable cover/index sheet."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 15

    title_font = Font(name="Arial", size=16, bold=True, color="1F4E79")
    sub_font = Font(name="Arial", size=10, color="404040")
    head_font = Font(name="Arial", size=10, bold=True, color=CLR_FONT_LIGHT)
    head_fill = _make_header_fill(CLR_HEADER_DARK)
    link_font = Font(name="Arial", size=10, color="2E75B6", underline="single")

    ws["B2"] = "NigeriaAgriScope — Power BI Data Mart"
    ws["B2"].font = title_font
    ws["B3"] = (
        "Author: Fidelis Akinbule  |  April 2026  |  Sources: FAOSTAT, NASA POWER, World Bank, USDA PSD"
    )
    ws["B3"].font = sub_font

    headers = ["Sheet", "Description", "Rows (approx)"]
    for c_idx, h in enumerate(headers, start=2):
        cell = ws.cell(row=5, column=c_idx)
        cell.value = h
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")

    sheet_index = [
        ("_COVER", "This index sheet", "—"),
        (
            "dim_zone_crop",
            "Master dimension: 42 zone-crop pairs (star-schema hub)",
            "42",
        ),
        (
            "fact_yield_history",
            "Historical yield, climate, fertilizer 2000–2023",
            "1,008",
        ),
        (
            "fact_yield_forecast",
            "XGBoost ML predictions vs actuals (2019–2023)",
            "~210",
        ),
        (
            "fact_production_forecast",
            "Prophet 3-crop national production 2000–2026",
            "~81",
        ),
        ("fact_input_requirements", "Fertilizer + seed + labour + cost per ha", "42"),
        (
            "fact_planting_calendar",
            "Optimal planting windows + reliability scores",
            "42",
        ),
        ("fact_operations_summary", "Key milestone months per zone-crop", "42"),
        ("fact_operations_detail", "Full week-by-week agronomic schedule", "~630"),
    ]

    data_font_cov = Font(name="Arial", size=9, color=CLR_FONT_DARK)
    alt = _make_alt_fill()
    for r_idx, (sheet, desc, rows) in enumerate(sheet_index, start=6):
        fill = alt if r_idx % 2 == 0 else None
        for c_idx, val in enumerate([sheet, desc, rows], start=2):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.font = link_font if c_idx == 2 else data_font_cov
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    ws["B16"] = "Power BI import guide: see README_PowerBI.md in the same folder."
    ws["B16"].font = Font(name="Arial", size=9, italic=True, color="606060")


# ── README writer ──────────────────────────────────────────────────────────

README_CONTENT = """# NigeriaAgriScope — Power BI Import Guide

**File:** `NigeriaAgriScope_PowerBI_DataMart.xlsx`
**Author:** Fidelis Akinbule | April 2026

---

## Step 1 — Import the workbook into Power BI Desktop

1. Open **Power BI Desktop**
2. Click **Home → Get Data → Excel Workbook**
3. Select `NigeriaAgriScope_PowerBI_DataMart.xlsx`
4. In the Navigator panel, tick **all 8 data sheets** (exclude `_COVER`):
   - `dim_zone_crop`
   - `fact_yield_history`
   - `fact_yield_forecast`
   - `fact_production_forecast`
   - `fact_input_requirements`
   - `fact_planting_calendar`
   - `fact_operations_summary`
   - `fact_operations_detail`
5. Click **Transform Data** → in Power Query, verify column types are correct
   (year = Whole Number, costs = Decimal, text = Text)
6. Click **Close & Apply**

---

## Step 2 — Build the star-schema relationships

Go to **Model view** (left sidebar icon). Create these relationships:

| From table (many side) | Column | To table (one side) | Column |
|---|---|---|---|
| fact_yield_history | zone_crop_key | dim_zone_crop | zone_crop_key |
| fact_yield_forecast | zone_crop_key | dim_zone_crop | zone_crop_key |
| fact_input_requirements | zone_crop_key | dim_zone_crop | zone_crop_key |
| fact_planting_calendar | zone_crop_key | dim_zone_crop | zone_crop_key |
| fact_operations_summary | zone_crop_key | dim_zone_crop | zone_crop_key |
| fact_operations_detail | zone_crop_key | dim_zone_crop | zone_crop_key |

`fact_production_forecast` is national-level (no zone) — connect on `crop` only:

| From table | Column | To table | Column |
|---|---|---|---|
| fact_production_forecast | crop | dim_zone_crop | crop |

Set all relationship cardinality to **Many-to-One (*)→(1)**.

---

## Step 3 — Suggested DAX measures

Create these in a dedicated `_Measures` table:

```dax
-- Yield gap as percentage of PDR yield
Yield Gap % =
DIVIDE(
    SUM(fact_input_requirements[yield_gap_at_pdr_hg_ha]),
    SUM(fact_yield_history[yield_hg_ha]),
    0
)

-- Fertilizer efficiency ratio (yield per kg fertilizer)
Fertilizer Efficiency =
DIVIDE(
    AVERAGE(fact_yield_history[yield_hg_ha]),
    AVERAGE(fact_yield_history[fertilizer_kg_ha]),
    0
)

-- Total input cost (Naira, millions)
Total Input Cost (₦M) =
DIVIDE(SUM(fact_input_requirements[total_input_cost_naira_per_ha]), 1000000)

-- Model accuracy (test set MAPE)
Model MAPE % =
AVERAGEX(
    fact_yield_forecast,
    fact_yield_forecast[abs_pct_error]
)

-- Forecast vs actual production growth
Production Growth % =
VAR actual_2023 =
    CALCULATE(SUM(fact_production_forecast[production_tonnes]),
              fact_production_forecast[year] = 2023,
              fact_production_forecast[type] = "actual")
VAR forecast_2026 =
    CALCULATE(SUM(fact_production_forecast[production_tonnes]),
              fact_production_forecast[year] = 2026,
              fact_production_forecast[type] = "forecast")
RETURN DIVIDE(forecast_2026 - actual_2023, actual_2023, 0)
```

---

## Step 4 — Suggested report pages

### Page 1 — National Overview
- Card: Total crops tracked, zones covered, years of data
- Line chart: `fact_yield_history` yield_hg_ha by year (all crops)
- Map visual: zone → production_tonnes (use Nigeria shapefile or built-in map)
- Slicer: crop (from dim_zone_crop)

### Page 2 — Yield Intelligence
- Scatter: actual vs predicted yield (fact_yield_forecast)
- Bar: SHAP feature importance (static image from chart01 — import as image)
- Line: yield trend 2000–2023 by zone (fact_yield_history filtered by crop slicer)

### Page 3 — Production Forecast
- Line + shaded band: production_tonnes + CI bounds (fact_production_forecast)
- Slicer: crop (Cassava / Oil palm fruit / Maize)
- Card: 2026 forecast vs 2023 actual growth %

### Page 4 — Input Planning
- Stacked bar: cost breakdown per crop (fact_input_requirements)
- Table: zone × crop with recommended fertilizer range + PDR
- Gauge: current_avg_kg_ha vs recommended_max_kg_ha (per selected zone-crop)

### Page 5 — Planting Calendar
- Matrix: zone (rows) × month (cols), coloured by planting month
- Bar: rainfall reliability score by zone
- Table: planting + harvest months with growing duration

### Page 6 — Operations Schedule (Drill-through)
- Table: fact_operations_detail filtered by selected zone-crop
- Set drill-through from Page 5 on zone_crop_key → this page
- Timeline visual: phase × week_offset coloured by phase category

---

## Data sources

| Source | Coverage | API |
|---|---|---|
| FAOSTAT QCL | Crop production 2000–2023 | fenixservices.fao.org |
| FAOSTAT RFN | Fertilizer by nutrient | fenixservices.fao.org |
| NASA POWER | Monthly climate 6 zones | power.larc.nasa.gov |
| World Bank WDI | Agricultural macro-indicators | api.worldbank.org |
| USDA PSD | Palm oil & cassava supply/demand | apps.fas.usda.gov |

**Model details:** XGBoost R²=0.991, MAE=2,900 hg/ha | Prophet 80% CI | PDR threshold=50 hg/ha per kg/ha
**Price benchmarks:** NBS Q3 2024 | ₦820/kg NPK 15:15:15 | ₦750/kg Urea | Exchange rate ₦1,580/USD
"""


# ── Entry point ────────────────────────────────────────────────────────────


def main() -> None:
    log.info("NigeriaAgriScope — Module 6: Power BI Data Mart Export")
    log.info("=" * 60)

    log.info("STEP 1 — Validate upstream sources")
    validate_sources()

    log.info("STEP 2 — Build all 8 tables")
    sheets: dict[str, tuple[pd.DataFrame, str, dict | None]] = {}
    for sheet_name, builder, header_colour, col_widths in SHEET_CONFIGS:
        df = builder()
        sheets[sheet_name] = (df, header_colour, col_widths)

    log.info("STEP 3 — Write and format Excel workbook")
    write_workbook(sheets)

    log.info("STEP 4 — Write Power BI import README")
    README_PATH.write_text(README_CONTENT, encoding="utf-8")
    log.info("README saved → %s", README_PATH.name)

    log.info("=" * 60)
    log.info("Module 6 — export_powerbi.py COMPLETE")
    log.info("  Output → module6_powerbi/outputs/")
    log.info("  NigeriaAgriScope_PowerBI_DataMart.xlsx")
    log.info("  README_PowerBI.md")
    log.info("")
    log.info("  Sheet summary:")
    for name, (df, _, _) in sheets.items():
        log.info("    %-30s  %d rows × %d cols", name, *df.shape)
    log.info("=" * 60)
    log.info(
        "Next: open README_PowerBI.md and follow Step 1 to import into Power BI Desktop."
    )


if __name__ == "__main__":
    main()
